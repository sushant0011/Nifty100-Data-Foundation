"""
exporter.py — Excel Export Engine for Nifty100 Screener & Peer Comparison.

Day 17: Generate screener_output.xlsx with 6 preset sheets, colour-coded cells.
Day 20: Generate peer_comparison.xlsx with 11 peer group sheets, percentile colours.
"""

import sqlite3
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.config import cfg
from src.utils.logger import logger
from src.screener.engine import ScreenerEngine


# ── Colour constants ───────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
GOLD   = PatternFill("solid", fgColor="FFD700")
HEADER = PatternFill("solid", fgColor="1F4E79")

HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT    = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

OUTPUT_DIR = cfg.OUTPUT_DIR

# KPI columns to include in exports
KPI_COLUMNS = [
    "company_id", "company_name", "year", "broad_sector",
    "net_profit_margin_pct", "operating_profit_margin_pct",
    "return_on_equity_pct", "return_on_capital_employed_pct",
    "return_on_assets_pct", "debt_to_equity", "interest_coverage",
    "icr_label", "net_debt_cr", "asset_turnover",
    "free_cash_flow_cr", "capex_intensity_pct", "capex_intensity_label",
    "fcf_conversion_rate_pct", "cfo_quality_score", "cfo_quality_label",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "composite_quality_score",
]


# ── Helpers ────────────────────────────────────────────────────────────────

def auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def write_header_row(ws, columns: list):
    """Write a styled header row."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN


def round_val(val):
    """Round floats to 2 decimal places for display."""
    if isinstance(val, float):
        return round(val, 2)
    return val


# ── Screener Excel Export ──────────────────────────────────────────────────

def colour_screener_cell(cell, col_name: str, value, filters: dict):
    """Green if the cell's metric passes the preset threshold, red otherwise."""
    threshold_map = {
        "return_on_equity_pct":       ("roe_min",              ">="),
        "debt_to_equity":             ("de_max",               "<="),
        "free_cash_flow_cr":          ("fcf_min",              ">="),
        "revenue_cagr_5yr":           ("revenue_cagr_5yr_min", ">="),
        "pat_cagr_5yr":               ("pat_cagr_5yr_min",     ">="),
        "operating_profit_margin_pct":("opm_min",              ">="),
        "interest_coverage":          ("icr_min",              ">="),
        "asset_turnover":             ("asset_turnover_min",   ">="),
        "net_profit_margin_pct":      ("net_profit_min",       ">="),
        "eps_cagr_5yr":               ("eps_cagr_min",         ">="),
    }
    if col_name not in threshold_map:
        return
    metric_key, op = threshold_map[col_name]
    if metric_key not in filters:
        return
    threshold = filters[metric_key]
    if value is None or (isinstance(value, float) and np.isnan(value)):
        cell.fill = RED
        return
    passes = (value >= threshold) if op == ">=" else (value <= threshold)
    cell.fill = GREEN if passes else RED


def export_screener_excel(output_path: Path = None):
    """Generate screener_output.xlsx with one sheet per preset."""
    if output_path is None:
        output_path = OUTPUT_DIR / "screener_output.xlsx"

    engine = ScreenerEngine()
    config = engine.config
    presets = config.get("presets", {})

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for preset_name, preset_info in presets.items():
        label = preset_info.get("label", preset_name)
        filters = preset_info.get("filters", {})

        result_df = engine.run_preset(preset_name)

        # Select available KPI columns
        available_cols = [c for c in KPI_COLUMNS if c in result_df.columns]
        export_df = result_df[available_cols].copy()

        ws = wb.create_sheet(title=label[:31])  # Excel sheet name limit 31 chars

        # Header
        write_header_row(ws, available_cols)

        # Data rows
        for row_idx, (_, data_row) in enumerate(export_df.iterrows(), start=2):
            for col_idx, col_name in enumerate(available_cols, start=1):
                val = data_row[col_name]
                val = round_val(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = CENTER_ALIGN
                colour_screener_cell(cell, col_name, val, filters)

        auto_width(ws)
        ws.freeze_panes = "A2"
        logger.info(f"  Sheet [{label}]: {len(export_df)} companies")

    wb.save(output_path)
    logger.success(f"screener_output.xlsx saved -> {output_path}")
    return output_path


# ── Peer Comparison Excel Export ───────────────────────────────────────────

def colour_percentile_cell(cell, percentile: float):
    """Green >= 75th, Yellow 25th-75th, Red <= 25th percentile."""
    if percentile is None or np.isnan(percentile):
        return
    if percentile >= 0.75:
        cell.fill = GREEN
    elif percentile >= 0.25:
        cell.fill = YELLOW
    else:
        cell.fill = RED


def export_peer_comparison_excel(output_path: Path = None):
    """Generate peer_comparison.xlsx with one sheet per peer group."""
    if output_path is None:
        output_path = OUTPUT_DIR / "peer_comparison.xlsx"

    conn = sqlite3.connect(cfg.DB_PATH)

    # Load peer percentiles
    try:
        peer_df = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
    except Exception as e:
        logger.error(f"peer_percentiles table not found: {e}")
        conn.close()
        return None

    if peer_df.empty:
        logger.warning("peer_percentiles table is empty — run PeerEngine first")
        conn.close()
        return None

    # Load company names
    companies = pd.read_sql_query("SELECT id as company_id, company_name FROM companies", conn)

    # Load latest ratios for metric values
    ratios = pd.read_sql_query("""
        SELECT fr.*
        FROM financial_ratios fr
        INNER JOIN (
            SELECT company_id, MAX(year) as max_year
            FROM financial_ratios GROUP BY company_id
        ) latest ON fr.company_id = latest.company_id AND fr.year = latest.max_year
    """, conn)
    conn.close()

    ratios = ratios.merge(companies, on="company_id", how="left")

    wb = Workbook()
    wb.remove(wb.active)

    peer_groups = peer_df["peer_group_name"].unique()

    for group_name in sorted(peer_groups):
        group_peer = peer_df[peer_df["peer_group_name"] == group_name]
        group_companies = group_peer["company_id"].unique()
        group_ratios = ratios[ratios["company_id"].isin(group_companies)].copy()

        if group_ratios.empty:
            continue

        # Pivot percentiles: company x metric
        pivot = group_peer.pivot_table(
            index="company_id",
            columns="metric",
            values="percentile_rank",
            aggfunc="first"
        ).reset_index()

        # Merge with ratios
        merged = group_ratios.merge(pivot, on="company_id", how="left", suffixes=("", "_pct"))

        # Columns: company_id, company_name, KPI cols, percentile rank cols
        metric_cols = [c for c in merged.columns if c in KPI_COLUMNS]
        pct_cols = [c for c in merged.columns if c in peer_df["metric"].unique()]

        all_cols = ["company_id", "company_name"] + metric_cols + [f"{m}_rank" for m in pct_cols]

        # Rename percentile cols
        rename_map = {m: f"{m}_rank" for m in pct_cols}
        merged = merged.rename(columns=rename_map)

        ws = wb.create_sheet(title=str(group_name)[:31])

        # Build column list
        display_cols = ["company_id", "company_name"]
        for c in KPI_COLUMNS:
            if c in merged.columns and c not in display_cols:
                display_cols.append(c)
        for m in pct_cols:
            rank_col = f"{m}_rank"
            if rank_col in merged.columns:
                display_cols.append(rank_col)

        display_cols = [c for c in display_cols if c in merged.columns]

        write_header_row(ws, display_cols)

        # Compute median row
        numeric_cols = merged[display_cols].select_dtypes(include=[np.number]).columns

        for row_idx, (_, data_row) in enumerate(merged.iterrows(), start=2):
            is_benchmark = False  # can be set per group if needed
            for col_idx, col_name in enumerate(display_cols, start=1):
                val = data_row.get(col_name)
                val = round_val(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = CENTER_ALIGN

                # Colour percentile rank cells
                if col_name.endswith("_rank"):
                    colour_percentile_cell(cell, val)

            if is_benchmark:
                for col_idx in range(1, len(display_cols) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = GOLD

        # Summary median row
        median_row_idx = len(merged) + 2
        ws.cell(row=median_row_idx, column=1, value="MEDIAN").font = BOLD_FONT
        for col_idx, col_name in enumerate(display_cols, start=1):
            if col_name in numeric_cols:
                median_val = round_val(merged[col_name].median())
                cell = ws.cell(row=median_row_idx, column=col_idx, value=median_val)
                cell.font = BOLD_FONT
                cell.alignment = CENTER_ALIGN

        auto_width(ws)
        ws.freeze_panes = "A2"
        logger.info(f"  Sheet [{group_name}]: {len(merged)} companies")

    wb.save(output_path)
    logger.success(f"peer_comparison.xlsx saved -> {output_path}")
    return output_path


if __name__ == "__main__":
    export_screener_excel()
    export_peer_comparison_excel()